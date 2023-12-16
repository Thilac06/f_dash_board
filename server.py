import sys
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QTableWidget, QTableWidgetItem
from PyQt5.QtGui import QColor
from openpyxl import load_workbook
import socket

host = '127.0.0.1'
port = 12345

        
server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)

        
server_socket.bind((host, port))

        
server_socket.listen(5)
print(f"Server listening on {host}:{port}")








class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1920, 1080)
        MainWindow.setStyleSheet("#main {\n"
            "  background-color: white;\n"
            "}\n"
            "\n"
            "#left {\n"
            "  background-color: #0f4d3e;\n"
            "}\n"
            "\n"
            "#table {\n"
            "  border: 1px solid black;\n"
            "}\n"
            "\n"
            "#geninfo, #Dash, #actual, #annual {\n"
            "  background-color: transparent;\n"
            "  font-family: \"Arial\", sans-serif;\n"
            "}\n"
            "\n"
            "#geninfo {\n"
            "  color: #fff;\n"
            "  font-size: 30px;\n"
            "}\n"
            "#Dash {\n"
            "  color: black;\n"
            "  font-size: 50px;\n"
            "}\n"
            "#actual, #annual {\n"
            "  color: #fff;\n"
            "  font-size: 30px;\n"
            "  text-align: left;\n"
            "}\n"
            "\n"
            "#actual:hover, #annual:hover {\n"
            "  background-color: gray;\n"
            "}\n"
        )

        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")

        self.main = QtWidgets.QWidget(self.centralwidget)
        self.main.setGeometry(QtCore.QRect(10, 10, 1920, 1080))
        self.main.setObjectName("main")

        self.left = QtWidgets.QWidget(self.main)
        self.left.setGeometry(QtCore.QRect(0, 0, 384, 1080))
        self.left.setObjectName("left")

        self.geninfo = QtWidgets.QComboBox(self.left)
        self.geninfo.setGeometry(QtCore.QRect(3, 200, 384, 41))
        self.geninfo.setObjectName("geninfo")
        self.geninfo.addItem("General Info A")
        self.geninfo.addItem("General Info B")
        self.geninfo.addItem("General Info C")

        self.tableView = QtWidgets.QTableWidget(self.main)
        self.tableView.setGeometry(QtCore.QRect(450, 161, 1380, 800))
        self.tableView.setObjectName("tableView")

        self.Dash = QtWidgets.QPushButton(self.left)
        self.Dash.setGeometry(QtCore.QRect(0, 27, 384, 51))
        self.Dash.setObjectName("Dash")

        self.actual = QtWidgets.QPushButton(self.left)
        self.actual.setGeometry(QtCore.QRect(3, 300, 384, 41))
        self.actual.setObjectName("actual")

        self.annual = QtWidgets.QPushButton(self.left)
        self.annual.setGeometry(QtCore.QRect(3, 400, 384, 41))
        self.annual.setObjectName("annual")

        self.pro = QtWidgets.QComboBox(self.main)
        self.pro.setGeometry(QtCore.QRect(450, 90, 201, 41))
        self.pro.setObjectName("pro")
        self.pro.addItem("PROVICE")
        self.pro.addItem("Central")
        self.pro.addItem("Eastern")
        self.pro.addItem("Northern")
        self.pro.addItem("Southern")
        self.pro.addItem("Western")
        self.pro.addItem("North_central")
        self.pro.addItem("Uva")
        self.pro.addItem("North_western")
        self.pro.addItem("Sabaragamuwa")
        self.pro.currentIndexChanged.connect(self.pro_changed)

        self.tla = QtWidgets.QComboBox(self.main)
        self.tla.setGeometry(QtCore.QRect(700, 90, 201, 41))
        self.tla.setObjectName("tla")
        self.tla.addItem("DISTRICT")

        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1561, 26))
        self.menubar.setObjectName("menubar")

        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        MainWindow.setCentralWidget(self.centralwidget)

        
        self.geninfo.currentIndexChanged.connect(self.updateMainBox)
        
    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.Dash.setText(_translate("MainWindow", "Dash Board"))
        self.actual.setText(_translate("MainWindow", "Actual"))
        self.annual.setText(_translate("MainWindow", "Annual"))

    def pro_changed(self):
        self.tla.clear()

        selected_pro = self.pro.currentText()
        if selected_pro == "PROVICE":
            self.tla.addItems(["DISTRICT"])
        elif selected_pro == "Central":
            self.tla.addItems(["Kandy", "Matale", "Nuwara-Eliya"])
        elif selected_pro == "Eastern":
            self.tla.addItems(["Ampara", "Batticaloa", "Trincomalee"])
        elif selected_pro == "Northern":
            self.tla.addItems(["Jaffna", "Killinochi", "Mannar", "Vavuniya", "Mulaithiwu"])
        elif selected_pro == "Southern":
            self.tla.addItems(["Galle", "Matara", "Hambantota"])
        elif selected_pro == "Western":
            self.tla.addItems(["Colombo", "Gampaha", "Kalutara"])
        elif selected_pro == "North_central":
            self.tla.addItems(["Anuradhapura", "Polonnaruwa"])
        elif selected_pro == "Uva":
            self.tla.addItems(["Badulla", "Moneragala"])
        elif selected_pro == "North_western":
            self.tla.addItems(["Kurunegala", "Puttalam"])
        elif selected_pro == "Sabaragamuwa":
            self.tla.addItems(["Kegalle", "Ratnapura"])

    def updateMainBox(self):
        index = self.geninfo.currentIndex()
        selected_item = self.geninfo.itemText(index)

        if selected_item == "General Info A":
            self.updateMainBoxForInfoA()
        elif selected_item == "General Info B":
            self.updateMainBoxForInfoB()
        elif selected_item == "General Info C":
            self.updateMainBoxForInfoC()

       
        server_socket.settimeout(1)  

        try:
            client_socket, client_address = server_socket.accept()
            print(f"Connection from {client_address}")

            data = client_socket.recv(1024)
            print(f"Received data: {data.decode('utf-8')}")

           
            client_socket.sendall(selected_item.encode('utf-8'))

            
        except socket.timeout:
            print("No incoming connections during the timeout period")

        
        server_socket.settimeout(None)
            

    def updateMainBoxForInfoA(self):
        self.updateMainBoxForInfo("A.xlsx")

    def updateMainBoxForInfoB(self):
        self.updateMainBoxForInfo("form_data_output.xlsx")

    def updateMainBoxForInfoC(self):
        self.updateMainBoxForInfo("rutput_data (1).xlsx")

    def updateMainBoxForInfo(self, file_path):
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active

            headers = [cell.value for cell in sheet[1]]

            self.tableView.setColumnCount(len(headers))
            self.tableView.setRowCount(sheet.max_row - 1)
            self.tableView.setHorizontalHeaderLabels(headers)

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                for col, value in enumerate(row):
                    item = QTableWidgetItem(str(value))
                    self.tableView.setItem(row_idx, col, item)

                    if col != 0:
                        is_numerical = isinstance(value, (int, float))
                        if is_numerical:
                            if value == 0.0:
                                item.setBackground(QColor(0, 0, 255))  # Blue
                            else:
                                item.setBackground(QColor(0, 255, 0))  # Green
                        else:
                            item.setBackground(QColor(255, 0, 0))  # Red

            self.displayRowSum(sheet)

        except Exception as e:
            print(f"Error loading Excel data: {e}")

    def displayRowSum(self, sheet):
        
        pass

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)

    excel_file_path = "A.xlsx"
    ui.updateMainBoxForInfo(excel_file_path)

    MainWindow.show()
    sys.exit(app.exec_())

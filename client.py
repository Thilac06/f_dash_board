import socket
import sys
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QTableWidget, QTableWidgetItem
from PyQt5.QtGui import QColor
from openpyxl import load_workbook
import threading

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1920, 1080)
        MainWindow.setStyleSheet("#main {\n"
            "  background-color: white;\n"
            "}\n"
            "\n"

            "\n"
            "#table {\n"
            "  border: 1px solid black;\n"
            "}\n"
            "\n"


        )

        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")

        self.main = QtWidgets.QWidget(self.centralwidget)
        self.main.setGeometry(QtCore.QRect(10, 10, 1920, 1080))
        self.main.setObjectName("main")





        self.tableView = QtWidgets.QTableWidget(self.main)
        self.tableView.setGeometry(QtCore.QRect(0,0,1920, 1080))
        self.tableView.setObjectName("tableView")






        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        MainWindow.setCentralWidget(self.centralwidget)

            

        

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))

    def updateMainBoxForInfo(self, file_path):
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active

            headers = [cell.value for cell in sheet[1]]

            self.tableView.setColumnCount(len(headers))
            self.tableView.setRowCount(sheet.max_row - 1)
            self.tableView.setHorizontalHeaderLabels(headers)

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                for col, value in enumerate(row):
                    item = QTableWidgetItem(str(value))
                    self.tableView.setItem(row_idx, col, item)

                    if col != 0:
                        is_numerical = isinstance(value, (int, float))
                        if is_numerical:
                            if value == 0.0:
                                item.setBackground(QColor(0, 0, 255))  # Blue
                            else:
                                item.setBackground(QColor(0, 255, 0))  # Green
                        else:
                            item.setBackground(QColor(255, 0, 0))  # Red

            self.displayRowSum(sheet)

        except Exception as e:
            print(f"Error loading Excel data: {e}")
    def displayRowSum(self, sheet):

        pass
def listen_to_server(ui, excel_file_paths):
    host = '127.0.0.1'
    port = 12345

    while True:
        client_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)

        try:
            client_socket.connect((host, port))
            print(f"Connected to {host}:{port}")

            message = "Hello, server!"
            client_socket.sendall(message.encode('utf-8'))
            print(f"Sent data: {message}")

            data = client_socket.recv(1024)
            print(f"Received echoed data: {data.decode('utf-8')}")

            x = data.decode('utf-8')
            print(f"x: {x}")

            if x in ["General Info A", "General Info B", "General Info C"]:
                ui.updateMainBoxForInfo(excel_file_paths[x])
            else:
                print("Invalid value of x")

            app.processEvents()

        except Exception as e:
            print(f"Error: {e}")
        finally:
            client_socket.close()

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)

    excel_file_paths = {
        "General Info A": "A.xlsx",
        "General Info B": "form_data_output.xlsx",
        "General Info C": "rutput_data (1).xlsx"
    }

    server_thread = threading.Thread(target=listen_to_server, args=(ui, excel_file_paths))
    server_thread.start()

    MainWindow.show()
    sys.exit(app.exec_())
